"""Integration tests for workbook parsing and projection reconstruction."""

from __future__ import annotations

import shutil
import sqlite3
import tempfile
import unittest
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from algobotdash.config import (
    ConfigurationError,
    ImportConfig,
    StrategyGroup,
    load_config,
)
from algobotdash.parser import _number, read_report
from algobotdash.service import ImportService
from tests.fixture_helpers import workbook


def config(source: Path) -> ImportConfig:
    """Build the standard configuration used by parser tests."""
    return ImportConfig(
        source_path=source,
        symbol_prefixes=(("WIN", "WIN"), ("WDO", "WDO")),
        strategy_groups=(
            StrategyGroup("Turtle", ("turtle",)),
            StrategyGroup("FVG", ("fvg",)),
        ),
    )


def remove_section_label(path: Path, section: str) -> None:
    """Remove a section marker from a workbook fixture."""
    book = load_workbook(path)
    sheet = book.active
    if sheet is None:
        raise RuntimeError("workbook fixture has no active worksheet")
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row, 1).value == section:
            sheet.cell(row, 1).value = None
            book.save(path)
            return
    raise AssertionError(f"section not found: {section}")


def truncate_positions_header(path: Path) -> None:
    """Make the positions header invalid in a workbook fixture."""
    book = load_workbook(path)
    sheet = book.active
    if sheet is None:
        raise RuntimeError("workbook fixture has no active worksheet")
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row, 1).value == "Posições":
            sheet.cell(row + 1, 13).value = None
            book.save(path)
            return
    raise AssertionError("section not found: Posições")


def append_rejected_transaction(path: Path) -> None:
    """Append a transaction with invalid required fields."""
    book = load_workbook(path)
    sheet = book.active
    if sheet is None:
        raise RuntimeError("workbook fixture has no active worksheet")
    sheet.append(
        (
            "2026.08.01 13:00:00", None, "WINQ26", "sell", "out", 1,
            130100, 1001, 0, 0, 0, 0, 10105, "invalid transaction",
        )
    )
    book.save(path)


def append_invalid_numeric_transaction(
    path: Path, column: int = 8, value: Any = "nan"
) -> None:
    """Append a transaction with a selected numeric value."""
    book = load_workbook(path)
    sheet = book.active
    if sheet is None:
        raise RuntimeError("workbook fixture has no active worksheet")
    values: list[Any] = [
        "2026.08.01 13:00:00", 104, "WINQ26", "sell", "out", 1,
        130100, 1001, 0, 0, 0, 0, 10105, "invalid numeric",
    ]
    values[column] = value
    sheet.append(tuple(values))
    book.save(path)


class ImportPipelineTests(unittest.TestCase):
    """Verify parsing, validation, and projection persistence."""
    tmp_path: Path = Path()

    def setUp(self) -> None:
        """Create an isolated temporary workspace for the test."""
        self.tmp_path = Path(tempfile.mkdtemp(prefix="algobotdash-tests-"))
        self.addCleanup(shutil.rmtree, self.tmp_path)

    def test_refresh_reconstructs_positions_orders_and_quality_counts(self) -> None:
        """Refresh should persist records and quality counters."""
        tmp_path = self.tmp_path
        source = tmp_path / "history.xlsx"
        database = tmp_path / "data" / "trades.sqlite"
        workbook(source)

        result = ImportService(config(source)).refresh(database)

        self.assertEqual(result.positions_created, 2)
        self.assertEqual(result.no_comment_count, 2)
        self.assertEqual(result.rejected_count, 0)
        connection = sqlite3.connect(database)
        self.assertEqual(
            connection.execute("select count(*) from positions").fetchone()[0], 2
        )
        self.assertEqual(
            connection.execute("select count(*) from orders").fetchone()[0], 2
        )
        self.assertEqual(
            connection.execute("select count(*) from transactions").fetchone()[0], 3
        )
        self.assertEqual(
            connection.execute(
                "select count(*) from transactions where direction = 'balance'"
            ).fetchone()[0],
            1,
        )
        self.assertEqual(
            connection.execute(
                "select volume_requested, volume_executed from positions "
                "where position_id = '1'"
            ).fetchone(),
            (2.0, 3.0),
        )
        self.assertEqual(
            connection.execute(
                "select volume_requested, volume_executed from orders "
                "where order_id = '1001'"
            ).fetchone(),
            (2.0, 2.0),
        )
        self.assertIsNone(
            connection.execute(
                "select strategy from positions where position_id = '1'"
            ).fetchone()[0]
        )
        self.assertIsNone(
            connection.execute(
                "select position_id from orders where order_id = '1001'"
            ).fetchone()[0]
        )
        self.assertEqual(
            connection.execute(
                "select strategy from orders where order_id = '1001'"
            ).fetchone()[0],
            "Turtle",
        )
        self.assertEqual(
            connection.execute(
                "select strategy from transactions where transaction_id = '101'"
            ).fetchone()[0],
            "Turtle",
        )
        self.assertIsNone(
            connection.execute(
                "select position_id from transactions where transaction_id = '101'"
            ).fetchone()[0]
        )
        self.assertEqual(
            connection.execute(
                "select symbol_family from positions where position_id = '1'"
            ).fetchone()[0],
            "WIN",
        )
        self.assertEqual(
            connection.execute(
                "select status from positions where position_id = '2'"
            ).fetchone()[0],
            "open",
        )
        self.assertEqual(
            connection.execute(
                "select no_comment_count, rejected_count from imports"
            ).fetchone(),
            (2, 0),
        )
        connection.close()

    def test_refresh_associates_position_with_matching_order_ticket(self) -> None:
        """A position inherits the strategy of its same-ticket opening order."""
        source = self.tmp_path / "history.xlsx"
        database = self.tmp_path / "data" / "trades.sqlite"
        workbook(source, legacy_report=True)

        result = ImportService(config(source)).refresh(database)

        self.assertEqual(result.no_comment_count, 1)
        connection = sqlite3.connect(database)
        self.assertEqual(
            connection.execute(
                "select strategy from positions where position_id = '1001'"
            ).fetchone()[0],
            "FVG",
        )
        self.assertIsNone(
            connection.execute(
                "select strategy from positions where position_id = '2'"
            ).fetchone()[0]
        )
        connection.close()

    def test_refresh_keeps_position_unassociated_when_ticket_symbol_differs(self) -> None:
        """A matching ticket cannot override a conflicting raw symbol."""
        source = self.tmp_path / "history.xlsx"
        database = self.tmp_path / "data" / "trades.sqlite"
        workbook(source, legacy_report=True)
        book = load_workbook(source)
        sheet = book.active
        if sheet is None:
            raise RuntimeError("workbook fixture has no active worksheet")
        sheet["C8"] = "WINV26"
        book.save(source)

        result = ImportService(config(source)).refresh(database)

        self.assertEqual(result.no_comment_count, 2)
        connection = sqlite3.connect(database)
        self.assertIsNone(
            connection.execute(
                "select strategy from positions where position_id = '1001'"
            ).fetchone()[0]
        )
        connection.close()


    def test_configuration_uses_longest_symbol_prefix(self) -> None:
        """Configuration should prefer the most specific symbol prefix."""
        configured = ImportConfig(
            source_path=Path("history.xlsx"),
            symbol_prefixes=(("W", "W"), ("WIN", "WIN")),
            strategy_groups=(),
        )

        self.assertEqual(configured.normalize_symbol("WIN$"), "WIN")


    def test_configuration_rejects_invalid_patterns(self) -> None:
        """Configuration should reject invalid regular expressions."""
        for patterns in ("turtle", [], [""], ["turtle", 1], ["("]):
            with self.subTest(patterns=patterns):
                path = self.tmp_path / "config.yml"
                path.write_text(
                    "source:\n  path: history.xlsx\nstrategies:\n  groups:\n"
                    f"    - name: Turtle\n      patterns: {patterns!r}\n",
                    encoding="utf-8",
                )
                expected = "pattern inválido" if patterns == ["("] else "patterns deve ser"
                with self.assertRaisesRegex(ConfigurationError, expected):
                    load_config(path)


    def test_configuration_rejects_empty_symbol_prefix(self) -> None:
        """Configuration should reject empty symbol prefixes."""
        path = self.tmp_path / "config.yml"
        path.write_text(
            "source:\n  path: history.xlsx\nsymbols:\n  prefixes:\n    '': Unknown\n",
            encoding="utf-8",
        )

        with self.assertRaisesRegex(ConfigurationError, "prefixos vazios"):
            load_config(path)


    def test_configuration_rejects_blank_symbol_family(self) -> None:
        """A symbol prefix must map to a non-blank analytical family."""
        path = self.tmp_path / "config.yml"
        path.write_text(
            "source:\n  path: history.xlsx\nsymbols:\n  prefixes:\n    WIN: '   '\n",
            encoding="utf-8",
        )

        with self.assertRaisesRegex(ConfigurationError, "famílias vazias"):
            load_config(path)


    def test_configuration_rejects_blank_strategy_name(self) -> None:
        """A strategy group must have a non-blank name."""
        path = self.tmp_path / "config.yml"
        path.write_text(
            "source:\n  path: history.xlsx\nstrategies:\n  groups:\n"
            "    - name: '   '\n      patterns: ['fvg']\n",
            encoding="utf-8",
        )

        with self.assertRaisesRegex(ConfigurationError, "name não vazio"):
            load_config(path)


    def test_configuration_rejects_non_mapping_root(self) -> None:
        """Configuration should require a mapping at the YAML root."""
        path = self.tmp_path / "config.yml"
        path.write_text("- invalid\n", encoding="utf-8")

        with self.assertRaisesRegex(ConfigurationError, "configuração deve ser um mapa"):
            load_config(path)


    def test_configuration_rejects_malformed_yaml(self) -> None:
        """Configuration should report malformed YAML clearly."""
        path = self.tmp_path / "malformed.yml"
        path.write_text("source:\n  path: [history.xlsx\n", encoding="utf-8")

        with self.assertRaisesRegex(ConfigurationError, "não foi possível interpretar"):
            load_config(path)


    def test_configuration_rejects_invalid_sections_and_source_path(self) -> None:
        """Configuration should validate section types and source paths."""
        cases = (
            ("source", "invalid", "source deve ser um mapa"),
            ("symbols", "invalid", "symbols deve ser um mapa"),
            ("strategies", "invalid", "strategies deve ser um mapa"),
            ("source", "{path: 42}", "source.path deve ser uma string"),
        )
        for section, value, message in cases:
            with self.subTest(section=section, value=value):
                path = self.tmp_path / f"{section}-{len(value)}.yml"
                prefix = "source:\n  path: history.xlsx\n"
                path.write_text(f"{prefix}{section}: {value}\n", encoding="utf-8")
                with self.assertRaisesRegex(ConfigurationError, message):
                    load_config(path)


    def test_number_rejects_non_finite_values(self) -> None:
        """Numeric parsing should reject non-finite values."""
        self.assertEqual(_number("1.5"), 1.5)
        self.assertIsNone(_number("nan"))
        self.assertIsNone(_number("inf"))
        self.assertIsNone(_number("-inf"))


    def test_transactions_reject_non_finite_numeric_values(self) -> None:
        """Transaction parsing should reject non-finite numeric values."""
        source = self.tmp_path / "history.xlsx"
        workbook(source)
        append_invalid_numeric_transaction(source)

        _, _, transactions, rejected, _ = read_report(source, config(source))

        self.assertEqual(len(transactions), 3)
        self.assertEqual(len(rejected), 1)
        self.assertEqual(rejected[0].raw_position_id, "104")
        self.assertEqual(rejected[0].reason, "valores numéricos inválidos em transação")


    def test_transactions_validate_each_numeric_column(self) -> None:
        """Every transaction numeric column should reject non-finite input."""
        for column in (5, 6, 8, 9, 10, 11, 12):
            with self.subTest(column=column):
                source = self.tmp_path / f"history-{column}.xlsx"
                workbook(source)
                append_invalid_numeric_transaction(source, column)

                _, _, transactions, rejected, _ = read_report(source, config(source))

                self.assertEqual(len(transactions), 3)
                self.assertEqual(len(rejected), 1)
                self.assertEqual(rejected[0].reason, "valores numéricos inválidos em transação")

    def test_transactions_treat_whitespace_numeric_values_as_blank(self) -> None:
        """Whitespace-only optional numeric cells should retain blank semantics."""
        for column in (5, 6, 8, 9, 10, 11, 12):
            with self.subTest(column=column):
                source = self.tmp_path / f"whitespace-{column}.xlsx"
                workbook(source)
                append_invalid_numeric_transaction(source, column, "   ")

                _, _, transactions, rejected, _ = read_report(source, config(source))

                self.assertEqual(len(transactions), 4)
                self.assertEqual(rejected, [])


    def test_new_source_preserves_import_history(self) -> None:
        """Refreshing a new source should preserve prior import history."""
        tmp_path = self.tmp_path
        source = tmp_path / "history.xlsx"
        changed_source = tmp_path / "history-changed.xlsx"
        database = tmp_path / "trades.sqlite"
        workbook(source)
        workbook(changed_source)
        changed_book = load_workbook(changed_source)
        changed_sheet = changed_book.active
        if changed_sheet is None:
            raise RuntimeError("workbook fixture has no active worksheet")
        changed_sheet["M3"] = 101
        changed_book.save(changed_source)

        ImportService(config(source)).refresh(database)
        ImportService(config(changed_source)).refresh(database)

        connection = sqlite3.connect(database)
        self.assertEqual(connection.execute("select count(*) from imports").fetchone()[0], 2)
        self.assertEqual(connection.execute("select count(*) from positions").fetchone()[0], 2)
        self.assertEqual(connection.execute("select count(*) from orders").fetchone()[0], 2)
        self.assertEqual(connection.execute("select count(*) from transactions").fetchone()[0], 3)
        self.assertEqual(connection.execute("select count(*) from rejected_rows").fetchone()[0], 0)
        self.assertEqual(
            connection.execute(
                "select strategy from transactions where transaction_id = '101'"
            ).fetchone()[0],
            "Turtle",
        )
        connection.close()


    def test_refresh_rejects_unsupported_schema_version(self) -> None:
        """Refresh should reject an unsupported database schema version."""
        tmp_path = self.tmp_path
        source = tmp_path / "history.xlsx"
        database = tmp_path / "trades.sqlite"
        workbook(source)
        connection = sqlite3.connect(database)
        connection.execute("create table schema_version (version integer primary key)")
        connection.execute("insert into schema_version values (99)")
        connection.execute("create table imports (id integer primary key)")
        connection.commit()
        connection.close()

        with self.assertRaisesRegex(ValueError, "versão de schema SQLite não suportada"):
            ImportService(config(source)).refresh(database)


    def test_refresh_rejects_unversioned_database(self) -> None:
        """Refresh should reject a database without schema metadata."""
        tmp_path = self.tmp_path
        source = tmp_path / "history.xlsx"
        database = tmp_path / "trades.sqlite"
        workbook(source)
        connection = sqlite3.connect(database)
        connection.execute("create table imports (positions_created integer)")
        connection.commit()
        connection.close()

        with self.assertRaisesRegex(ValueError, "schema_version ausente"):
            ImportService(config(source)).refresh(database)


    def test_refresh_rejects_malformed_report_without_mutating_projection(self) -> None:
        """Malformed reports should not replace an existing projection."""
        tmp_path = self.tmp_path
        source = tmp_path / "history.xlsx"
        database = tmp_path / "trades.sqlite"
        workbook(source)
        ImportService(config(source)).refresh(database)

        malformed_sources = {
            "missing_positions": ("missing_positions.xlsx", "Posições", False),
            "missing_orders": ("missing_orders.xlsx", "Ordens", False),
            "missing_transactions": ("missing_transactions.xlsx", "Transações", False),
            "truncated_header": ("truncated_header.xlsx", "", True),
        }
        for name, (filename, section, truncated) in malformed_sources.items():
            with self.subTest(name=name):
                malformed = tmp_path / filename
                workbook(malformed)
                if truncated:
                    truncate_positions_header(malformed)
                else:
                    remove_section_label(malformed, section)

                with self.assertRaises(ValueError):
                    ImportService(config(malformed)).refresh(database)

                connection = sqlite3.connect(database)
                self.assertEqual(
                    connection.execute("select count(*) from positions").fetchone()[0], 2
                )
                self.assertEqual(
                    connection.execute("select count(*) from orders").fetchone()[0], 2
                )
                self.assertEqual(
                    connection.execute("select count(*) from transactions").fetchone()[0], 3
                )
                self.assertEqual(
                    connection.execute("select count(*) from rejected_rows").fetchone()[0], 0
                )
                self.assertEqual(
                    connection.execute(
                        "select strategy from orders where order_id = '1001'"
                    ).fetchone()[0],
                    "Turtle",
                )
                self.assertEqual(
                    connection.execute(
                        "select strategy from transactions where transaction_id = '101'"
                    ).fetchone()[0],
                    "Turtle",
                )
                connection.close()


    def test_refresh_is_idempotent_by_replacing_projection(self) -> None:
        """Refreshing the same source should replace, not duplicate, records."""
        tmp_path = self.tmp_path
        source = tmp_path / "history.xlsx"
        database = tmp_path / "trades.sqlite"
        workbook(source)
        append_rejected_transaction(source)
        service = ImportService(config(source))

        first = service.refresh(database)
        second = service.refresh(database)

        self.assertEqual(first.source_hash, second.source_hash)
        connection = sqlite3.connect(database)
        self.assertEqual(connection.execute("select count(*) from imports").fetchone()[0], 1)
        self.assertEqual(connection.execute("select count(*) from positions").fetchone()[0], 2)
        self.assertEqual(connection.execute("select count(*) from orders").fetchone()[0], 2)
        self.assertEqual(connection.execute("select count(*) from transactions").fetchone()[0], 3)
        self.assertEqual(connection.execute("select count(*) from rejected_rows").fetchone()[0], 1)
        self.assertEqual(
            connection.execute("select reason from rejected_rows").fetchone()[0],
            "campos obrigatórios inválidos em transação",
        )
        self.assertEqual(
            connection.execute(
                "select strategy from orders where order_id = '1001'"
            ).fetchone()[0],
            "Turtle",
        )
        self.assertEqual(
            connection.execute(
                "select strategy from transactions where transaction_id = '101'"
            ).fetchone()[0],
            "Turtle",
        )
        connection.close()


    def test_ambiguous_comment_aborts_without_publishing_or_mutating_existing_projection(
        self,
    ) -> None:
        """Ambiguous strategy comments should abort without publishing changes."""
        tmp_path = self.tmp_path
        source = tmp_path / "history.xlsx"
        bad_source = tmp_path / "bad.xlsx"
        database = tmp_path / "trades.sqlite"
        workbook(source)
        workbook(bad_source, ambiguous=True)
        ImportService(config(source)).refresh(database)

        with self.assertRaisesRegex(ConfigurationError, "múltiplos grupos"):
            ImportService(config(bad_source)).refresh(database)

        connection = sqlite3.connect(database)
        self.assertEqual(connection.execute("select count(*) from positions").fetchone()[0], 2)
        connection.close()
