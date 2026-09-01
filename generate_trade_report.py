"""Generate the initial static trade report from the history workbook."""

# pyright: reportOptionalOperand=false
import html
import json
import math
import os
import random
import re
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path
from typing import TypedDict
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

from algobotdash.config import load_config

CONFIG_PATH = Path(os.getenv("ALGOBOTDASH_CONFIG", "config/config.yaml"))
OUT = Path("reports")
OUT.mkdir(exist_ok=True)
REPORT_TZ = ZoneInfo("America/Bahia")
START = datetime(2026, 4, 1, tzinfo=REPORT_TZ)
SEED = 20260829
COLORS = {
    "BIT FVG": "#f59e0b",
    "BIT Turtle": "#8b5cf6",
    "WDO FVG": "#06b6d4",
    "WDO RadarWDO": "#ef4444",
    "WDO Turtle": "#22c55e",
    "WIN EngulfPattern": "#f97316",
    "WIN FVG": "#3b82f6",
    "WIN RadarWIN": "#e11d48",
    "WIN Soberano": "#a855f7",
    "WIN Turtle": "#14b8a6",
}
class Trade(TypedDict):
    """Trade fields used to build report metrics."""

    entry: datetime
    symbol: str
    pnl: float
    strategy: str


class Metrics(TypedDict):
    """Summary metrics for a trade series."""

    n: int
    net: float
    pf: float | None
    wr: float | None
    pay: float | None
    exp: float
    dd: float
    sh: float | None
    so: float | None
    dur: int


class Bootstrap(TypedDict):
    """Bootstrap confidence intervals for a trade series."""

    net: tuple[float, float]
    exp: tuple[float, float]
    pf: tuple[float, float] | None
    p: float


def date(v: object) -> datetime | None:
    """Parse a report timestamp and attach the report timezone."""
    if isinstance(v, datetime):
        return v if v.tzinfo else v.replace(tzinfo=REPORT_TZ)
    for fmt in ("%Y.%m.%d %H:%M:%S", "%Y.%m.%d %H:%M"):
        try:
            return datetime.strptime(str(v), fmt).replace(tzinfo=REPORT_TZ)
        except ValueError:
            continue
    return None
def num(v: object) -> float | None:
    """Parse a finite floating-point value."""
    try:
        parsed = float(str(v))
        return parsed if math.isfinite(parsed) else None
    except (TypeError, ValueError):
        return None
def sym(v: object) -> str | None:
    """Normalize a symbol to its supported market family."""
    value = str(v or "").upper()
    return next((item for item in ("WIN", "WDO", "BIT") if value.startswith(item)), None)


def strat(v: object) -> str | None:
    """Normalize a strategy comment to a report strategy name."""
    value = str(v or "").lower()
    for pattern, name in (
        ("turtle", "Turtle"),
        ("fvg", "FVG"),
        ("radarwdo", "RadarWDO"),
        ("radarwin", "RadarWIN"),
        ("soberano", "Soberano"),
        ("engulf", "EngulfPattern"),
    ):
        if pattern in value:
            return name
    return None


def key(t: Trade) -> str:
    """Build a display key from a trade's symbol and strategy."""
    return f"{t['symbol']} {t['strategy']}"


def fbr(v: float | None, d: int = 0) -> str:
    """Format a number using Brazilian decimal and thousands separators."""
    if v is None:
        return "—"
    return f"{v:,.{d}f}".replace(",", "X").replace(".", ",").replace("X", ".")


def money(v: float) -> str:
    """Format a value as Brazilian currency."""
    return ("-" if v < 0 else "") + "R$ " + fbr(abs(v), 0)


def pct(v: float | None):
    """Format a ratio as a percentage."""
    return "—" if v is None else f"{fbr(v * 100, 1)}%"


def metric_number(value: object, default: float = 0.0) -> float:
    """Convert a metric value to float, using a safe default."""
    return float(value) if isinstance(value, (int, float)) else default


def m(xs: list[float]) -> Metrics:
    """Calculate summary metrics for a list of PnL values."""
    n = len(xs)
    if not n:
        return {
            "n": 0, "net": 0.0, "pf": None, "wr": None, "pay": None,
            "exp": 0.0, "dd": 0.0, "sh": None, "so": None, "dur": 0,
        }
    wins = [x for x in xs if x > 0]
    losses = [x for x in xs if x < 0]
    gross_profit = sum(wins)
    gross_loss = -sum(losses)
    avg = sum(xs) / n
    equity = peak = drawdown = duration = current_duration = 0
    for x in xs:
        equity += x
        if equity >= peak:
            peak = equity
            current_duration = 0
        else:
            current_duration += 1
            duration = max(duration, current_duration)
            drawdown = max(drawdown, peak - equity)
    standard_deviation = math.sqrt(
        sum((x - avg) ** 2 for x in xs) / (n - 1)
    ) if n > 1 else 0
    downside = math.sqrt(sum(min(0, x) ** 2 for x in xs) / n)
    return {
        "n": n,
        "net": sum(xs),
        "pf": gross_profit / gross_loss if gross_loss else (99 if gross_profit else None),
        "wr": len(wins) / n,
        "pay": (sum(wins) / len(wins)) / (gross_loss / len(losses))
        if wins and losses else None,
        "exp": avg,
        "dd": drawdown,
        "sh": avg / standard_deviation if standard_deviation else None,
        "so": avg / downside if downside else None,
        "dur": duration,
    }


def bs(xs: list[float], reps: int = 2000) -> Bootstrap | None:
    """Estimate bootstrap confidence intervals for PnL metrics."""
    if len(xs) < 10:
        return None
    rng = random.Random(SEED + len(xs))
    n = len(xs)
    outcomes = [m([xs[rng.randrange(n)] for _ in range(n)]) for _ in range(reps)]

    def ci(k: str) -> tuple[float, float]:
        values = sorted(x[k] for x in outcomes if x[k] is not None)
        return values[int(0.025 * len(values))], values[int(0.975 * len(values)) - 1]

    def optional_ci(k: str) -> tuple[float, float] | None:
        values = sorted(x[k] for x in outcomes if x[k] is not None)
        return (
            (values[int(0.025 * len(values))], values[int(0.975 * len(values)) - 1])
            if values else None
        )

    return {
        "net": ci("net"),
        "exp": ci("exp"),
        "pf": optional_ci("pf"),
        "p": sum(x["pf"] > 1 for x in outcomes if x["pf"] is not None)
        / len(outcomes),
    }


def mc(xs: list[float], reps: int = 2500) -> dict[str, float] | None:
    """Estimate drawdown threshold probabilities by Monte Carlo."""
    if len(xs) < 10:
        return None
    rng = random.Random(SEED + len(xs) * 3)
    n = len(xs)
    drawdowns = sorted(
        m([xs[rng.randrange(n)] for _ in range(n)])['dd'] for _ in range(reps)
    )
    return {
        "p1": sum(x >= 1000 for x in drawdowns) / reps,
        "p2": sum(x >= 2000 for x in drawdowns) / reps,
        "p95": drawdowns[int(0.95 * reps)],
    }


def svg_line(series: list[float], title: str, color: str = "#05ad98", h: int = 240) -> str:
    """Render one metric series as an inline SVG chart."""
    w = 920
    p = 36
    vals = series or [0]
    lo = min(vals)
    hi = max(vals)
    span = max(1, hi - lo)
    pts = []
    pts.extend(
        f"{p + i * (w - 2 * p) / max(1, len(vals) - 1):.1f},"
        f"{h - p - (v - lo) * (h - 2 * p) / span:.1f}"
        for i, v in enumerate(vals)
    )
    grid = "".join(
        f'<line x1="{p}" x2="{w-p}" y1="{p+i*(h-2*p)/4}" '
        f'y2="{p+i*(h-2*p)/4}" class="grid"/>'
        for i in range(5)
    )
    points = " ".join(pts)
    return (
        f'<svg viewBox="0 0 {w} {h}"><text x="{p}" y="18" '
        f'class="svgtitle">{title}</text>{grid}<polyline points="{points}" '
        f'class="line" style="stroke:{color}"/><text x="{p}" y="{h-8}" '
        f'class="axis">{money(lo)}</text><text x="{w-p-90}" y="{h-8}" '
        f'class="axis">{money(hi)}</text></svg>'
    )
# This chart assembles one SVG from several independently formatted fragments.
# pylint: disable=too-many-locals
def svg_multi(
    series: dict[str, list[float]],
    title: str,
    start_label: str = "",
    end_label: str = "",
) -> str:
    """Render several metric series as one inline SVG chart."""
    w, h, p = 920, 300, 64
    all_values = [x for values in series.values() for x in values] or [0]
    lo = min(all_values)
    hi = max(all_values)
    span = max(1, hi - lo)
    grid = "".join(
        f'<line x1="{p}" x2="{w-p}" y1="{p+i*(h-2*p)/4}" '
        f'y2="{p+i*(h-2*p)/4}" class="grid"/>'
        f'<text x="4" y="{p+i*(h-2*p)/4+4:.1f}" class="axis">'
        f'{money(hi-i*span/4)}</text>'
        for i in range(5)
    )
    lines = []
    legend = []
    for name, vals in series.items():
        points = " ".join(
            f"{p+i*(w-2*p)/max(1,len(vals)-1):.1f},"
            f"{h-p-(v-lo)*(h-2*p)/span:.1f}"
            for i, v in enumerate(vals)
        )
        color = COLORS.get(name, "#ffffff")
        lines.append(
            f'<polyline points="{points}" class="line" style="stroke:{color}"/>'
        )
        legend.append(
            f'<span class="legend"><i style="background:{color}"></i>'
            f'{html.escape(name)}</span>'
        )
    return (
        f'<div class="legendbox">{"".join(legend)}</div><svg '
        f'viewBox="0 0 {w} {h}"><text x="{p}" y="18" '
        f'class="svgtitle">{title}</text>{grid}{"".join(lines)}'
        f'<text x="{p}" y="{h-8}" class="axis">{start_label}</text>'
        f'<text x="{w-p-75}" y="{h-8}" class="axis">{end_label}</text></svg>'
    )


# This chart keeps dimensions and bar geometry together for readable markup.
# pylint: disable=too-many-locals
def svg_bars(labels: list[str], vals: list[float], title: str) -> str:
    """Render labeled values as an inline SVG bar chart."""
    w, h, p = 920, 250, 38
    scale = max(1, max(abs(x) for x in vals))
    zero = h / 2
    bw = (w - 2 * p) / max(1, len(vals)) * 0.62
    bars = []
    for i, (label, value) in enumerate(zip(labels, vals)):
        x = p + i * (w - 2 * p) / len(vals) + (w - 2 * p) / len(vals) * 0.19
        bar_height = abs(value) / scale * (h * 0.36)
        y = zero - bar_height if value >= 0 else zero
        color = "#05ad98" if value >= 0 else "#e85d6c"
        bars.append(
            f'<rect x="{x:.1f}" y="{y:.1f}" width="{bw:.1f}" '
            f'height="{bar_height:.1f}" fill="{color}"/>'
            f'<text x="{x:.1f}" y="{h-12}" class="axis">{label}</text>'
        )
    return (
        f'<svg viewBox="0 0 {w} {h}"><text x="{p}" y="18" '
        f'class="svgtitle">{title}</text><line x1="{p}" x2="{w-p}" '
        f'y1="{zero}" y2="{zero}" class="grid"/>{"".join(bars)}</svg>'
    )
# This is the legacy report's single orchestration boundary.
# pylint: disable=too-many-locals,too-many-branches,too-many-statements
def main():
    """Build and write the static HTML trade report."""
    source = load_config(CONFIG_PATH).source_path
    OUT.mkdir(exist_ok=True)
    book = load_workbook(source, read_only=True, data_only=True)
    sheet = book.active
    if sheet is None:
        raise RuntimeError(f"workbook has no active worksheet: {source}")
    rows = list(sheet.iter_rows(values_only=True))
    secs = {
        str(r[0]).strip(): i
        for i, r in enumerate(rows)
        if r
        and isinstance(r[0], str)
        and str(r[0]).strip() in {"Posições", "Ordens", "Transações"}
    }
    order_start = secs["Ordens"] + 2
    transaction_start = secs["Transações"]
    comments = {
        str(row[1]): str(row[11] or "")
        for row in rows[order_start:transaction_start]
        if len(row) > 11 and row[1]
    }
    trades: list[Trade] = []
    position_start = secs["Posições"] + 2
    order_end = secs["Ordens"]
    for row in rows[position_start:order_end]:
        entry = date(row[0]) if len(row) > 0 else None
        symbol = sym(row[2]) if len(row) > 2 else None
        pnl = num(row[12]) if len(row) > 12 else None
        strategy = strat(comments.get(str(row[1]), "")) if len(row) > 1 else None
        if entry is None or symbol is None or pnl is None or strategy is None or entry < START:
            continue
        if (
            symbol == "WIN"
            and strategy == "Turtle"
            and entry < datetime(2026, 6, 26, tzinfo=REPORT_TZ)
        ):
            continue
        trades.append({"entry": entry, "symbol": symbol, "pnl": pnl, "strategy": strategy})
    trades.sort(key=lambda trade: trade["entry"])
    if not trades:
        raise RuntimeError("nenhuma operação encontrada após os filtros do relatório")
    end = max(trade["entry"] for trade in trades)
    c30 = end - timedelta(days=30)
    week_start = end.date() - timedelta(days=end.weekday())
    groups: defaultdict[str, list[Trade]] = defaultdict(list)
    for trade in trades:
        groups[key(trade)].append(trade)
    keys = sorted(groups)
    full = m([trade["pnl"] for trade in trades])

    def sub(dataset: list[Trade], period: str) -> list[Trade]:
        """Return trades in one of the report's rolling periods."""
        return [
            trade
            for trade in dataset
            if (
                trade["entry"] >= c30
                if period == "30"
                else trade["entry"].date() >= week_start
            )
        ]

    daily = defaultdict(lambda: defaultdict(float))
    for trade in trades:
        daily[trade["entry"].date()][key(trade)] += trade["pnl"]
    days = sorted(daily)
    port = [sum(daily[day].values()) for day in days]
    eq = []
    total = 0
    for value in port:
        total += value
        eq.append(total)
    strategy_equity = {}
    for k in keys:
        running = 0
        values = []
        for day in days:
            running += daily[day][k]
            values.append(running)
        strategy_equity[k] = values
    dd = []
    peak = 0
    for value in eq:
        peak = max(peak, value)
        dd.append(value - peak)
    months = defaultdict(list)
    for trade in trades:
        months[trade["entry"].strftime("%Y-%m")].append(trade["pnl"])

    def corr(a: list[float], b: list[float]) -> float:
        """Calculate correlation between two daily PnL series."""
        mean_a = sum(a) / len(a)
        mean_b = sum(b) / len(b)
        variance_a = sum((x - mean_a) ** 2 for x in a)
        variance_b = sum((x - mean_b) ** 2 for x in b)
        covariance = sum(
            (x - mean_a) * (y - mean_b) for x, y in zip(a, b)
        )
        return covariance / math.sqrt(variance_a * variance_b) \
            if variance_a and variance_b else 0

    cm = [
        [corr([daily[d][a] for d in days], [daily[d][b] for d in days]) for b in keys]
        for a in keys
    ]
    info = []
    for k in keys:
        values = [trade["pnl"] for trade in groups[k]]
        all_time = m(values)
        recent_30 = m([trade["pnl"] for trade in sub(groups[k], "30")])
        recent_week = m([trade["pnl"] for trade in sub(groups[k], "week")])
        bootstrap = bs(values)
        tail = mc(values)
        pf20 = m(values[-20:])["pf"] if len(values) >= 20 else None
        if len(values) < 20:
            status, css_class = "Sombra", "warn"
        elif metric_number(all_time["pf"]) < 0.9 or (
            bootstrap and bootstrap["p"] < 0.4
        ):
            status, css_class = "Reduzir/Pausar", "bad"
        elif metric_number(all_time["pf"]) < 1 or (
            pf20 is not None and pf20 < 0.9
        ):
            status, css_class = "Sem aumento", "warn"
        else:
            status, css_class = "Continuar", "good"
        first = min(trade["entry"] for trade in groups[k])
        info.append((k, all_time, recent_30, recent_week, bootstrap, tail,
                     pf20, status, css_class, first))
    info.sort(key=lambda item: (-1 if item[1]["pf"] is None else -item[1]["pf"], item[0]))
    loo = []
    for k in keys:
        remaining = [sum(daily[d].values()) - daily[d][k] for d in days]
        without = m(remaining)
        loo.append((k, full["dd"] - without["dd"], without["dd"]))
    fixed = {"RadarWIN", "RadarWDO", "Soberano"}
    sizing = []
    for budget in (0.016, 0.017, 0.020, 0.030, 0.050):
        scaled = [
            trade["pnl"] * (1 if trade["strategy"] in fixed else budget / 0.016)
            for trade in trades
        ]
        sizing.append((budget, m(scaled)))

    def tr(c: list[str], tag: str = "td") -> str:
        """Render a table row."""
        cells = "".join(f"<{tag}>{cell}</{tag}>" for cell in c)
        return f"<tr>{cells}</tr>"

    perf = "".join(
        tr([
            f'<span class="dot" style="background:{COLORS[k]}"></span>{k}',
            first.strftime("%d/%m/%Y"),
            a["n"],
            f'<span class="{"pos" if a["net"] >= 0 else "neg"}">{money(a["net"])}</span>',
            fbr(a["pf"], 2), pct(a["wr"]), money(a["dd"]),
            fbr(a["sh"], 2), fbr(a["so"], 2),
            f'<span class="badge {cl}">{st}</span>',
        ])
        for k, a, _, _, _, _, _, st, cl, first in info
    )
    robust_info = sorted(
        info, key=lambda item: (item[4] is None, -(item[4]["p"] if item[4] else -1))
    )
    robust = "".join(
        tr([
            f'<span class="dot" style="background:{COLORS[k]}"></span>{k}',
            f'[{money(b["net"][0])}; {money(b["net"][1])}]' if b else "Amostra insuficiente",
            f'[{money(b["exp"][0])}; {money(b["exp"][1])}]' if b else "—",
            pct(b["p"]) if b else "—",
            money(tail["p95"]) if tail else "—",
            pct(tail["p1"]) if tail else "—",
            pct(tail["p2"]) if tail else "—",
        ])
        for k, _, _, _, b, tail, *_ in robust_info
    )

    def period_table(label: str, dataset) -> str:
        """Render the metrics table for one time period."""
        body = "".join(
            tr([
                f'<span class="dot" style="background:{COLORS[k]}"></span>{k}',
                metrics["n"], money(metrics["net"]), fbr(metrics["pf"], 2),
                pct(metrics["wr"]), fbr(metrics["pay"], 2), money(metrics["dd"]),
                fbr(metrics["sh"], 2), fbr(metrics["so"], 2),
            ])
            for k, a, r30, rw, *_ in info
            for metrics in [dataset(a, r30, rw)]
        )
        headings = [
            "Estratégia", "Operações", "P&L", "PF", "Tx. acerto",
            "Payoff", "DD", "Sharpe", "Sortino",
        ]
        return (
            f'<div class="period"><h3>{label}</h3><div class="scroll">'
            f'<table><thead>{tr(headings, "th")}</thead><tbody>{body}'
            "</tbody></table></div></div>"
        )

    temporal = (
        period_table("Desde o início operacional", lambda a, r, w: a)
        + period_table("Últimos 30 dias", lambda a, r, w: r)
        + period_table("Semana corrente", lambda a, r, w: w)
    )
    heat = ''.join(
        f'<tr><th>{html.escape(a)}</th>'
        + ''.join(
            f'<td class="heat" style="--v:{(v+1)/2:.3f}">{v:.2f}</td>'
            for v in row
        )
        + '</tr>'
        for a, row in zip(keys, cm)
    )
    sizetable = "".join(
        tr([pct(budget), money(metrics["net"]), fbr(metrics["pf"], 2),
            money(metrics["dd"]), money(metrics["exp"])])
        for budget, metrics in sizing
    )
    looh = "".join(
        tr([k, money(delta), money(xdd)])
        for k, delta, xdd in sorted(loo, key=lambda item: item[1], reverse=True)
    )
    monthly = sorted(months.items())
    monthlyh = "".join(
        tr([month, m(values)["n"], money(m(values)["net"]),
            fbr(m(values)["pf"], 2), money(m(values)["dd"])]
        )
        for month, values in monthly
    )
    strong = sorted(
        ((abs(cm[i][j]), keys[i], keys[j], cm[i][j])
         for i in range(len(keys)) for j in range(i + 1, len(keys))),
        reverse=True,
    )[:3]
    strongtxt = "; ".join(f"{a} x {b}: {value:.2f}" for _, a, b, value in strong)
    recent30 = m([trade["pnl"] for trade in trades if trade["entry"] >= c30])
    recentwk = m([trade["pnl"] for trade in trades if trade["entry"].date() >= week_start])
    title = "Avaliação Quantitativa de Estratégias"
    now = datetime.now(REPORT_TZ).strftime("%d/%m/%Y %H:%M")
    # The legacy report intentionally keeps its static HTML/CSS minified.
    # pylint: disable=line-too-long
    doc=f'''<!doctype html><html lang="pt-BR" data-theme="dark"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>{title}</title><meta name="report-type" content="research"><meta name="report-project" content="Trade"><meta name="report-date" content="{end:%Y-%m-%d}"><script id="report-meta" type="application/json">{json.dumps({'title':title,'type':'research','project':'Trade','date':str(end.date())})}</script><style>
:root{{--bg:#0a0f14;--panel:#0f1519;--hover:#151c22;--border:#24303a;--text:#e5edf4;--muted:#81909c;--pos:#05ad98;--neg:#e85d6c;--warn:#f5a623}}[data-theme=light]{{--bg:#fafafa;--panel:#fff;--hover:#eef2f4;--border:#ccd5da;--text:#0d1720;--muted:#5f6d77;--pos:#047e70;--neg:#c6283f;--warn:#a96800}}*{{box-sizing:border-box}}body{{margin:0;background:var(--bg);color:var(--text);font:14px Inter,Arial,sans-serif}}.shell{{max-width:1440px;margin:auto;padding:26px}}header{{display:flex;justify-content:space-between;gap:16px;align-items:flex-start;border-bottom:1px solid var(--border);padding-bottom:18px;margin-bottom:18px}}h1{{font-size:20px;letter-spacing:.07em;text-transform:uppercase;margin:0}}h3{{font-size:13px;margin:0 0 10px}}.sub,.muted{{color:var(--muted);font-size:12px}}button{{background:transparent;border:1px solid var(--border);color:var(--text);padding:7px 10px;cursor:pointer}}.metrics{{display:grid;grid-template-columns:repeat(5,1fr);gap:1px;background:var(--border);border:1px solid var(--border);margin-bottom:16px}}.metric{{background:var(--panel);padding:15px}}.metric label,.panelhead{{display:block;color:var(--muted);font:10px ui-monospace,monospace;letter-spacing:.13em;text-transform:uppercase}}.metric strong{{display:block;font:500 24px ui-monospace,monospace;margin:7px 0 3px}}.grid{{display:grid;grid-template-columns:1.55fr 1fr;gap:16px}}.panel{{background:var(--panel);border:1px solid var(--border);margin-bottom:16px}}.panelhead{{padding:11px 14px;border-bottom:1px solid var(--border)}}.body{{padding:14px}}.callout{{padding:14px;border-left:3px solid var(--warn);background:var(--hover);margin-bottom:12px}}.callout.good{{border-color:var(--pos)}}.callout.bad{{border-color:var(--neg)}}.callout p{{margin:4px 0 0}}table{{width:100%;border-collapse:collapse;font:12px ui-monospace,monospace}}th{{font-size:10px;color:var(--muted);letter-spacing:.06em;text-transform:uppercase;text-align:left}}th,td{{padding:9px 10px;border-bottom:1px solid var(--border);white-space:nowrap}}tr:hover{{background:var(--hover)}}.scroll{{overflow:auto}}.pos{{color:var(--pos)}}.neg{{color:var(--neg)}}.badge{{font:10px ui-monospace,monospace;border:1px solid var(--border);padding:3px 5px}}.badge.good{{color:var(--pos);border-color:var(--pos)}}.badge.bad{{color:var(--neg);border-color:var(--neg)}}.badge.warn{{color:var(--warn);border-color:var(--warn)}}svg{{width:100%;height:auto;background:#0a0f14}}.grid line,.grid{{stroke:#26323b;stroke-dasharray:2 3}}.line{{fill:none;stroke-width:2;vector-effect:non-scaling-stroke}}.svgtitle,.axis{{font:10px ui-monospace,monospace;fill:#81909c}}.heat{{background:color-mix(in srgb,var(--pos) calc(var(--v)*65%),var(--neg));color:#061015;text-align:center}}details{{border-top:1px solid var(--border);padding:12px 0}}summary{{cursor:pointer;font-weight:600}}footer{{color:var(--muted);font-size:11px;padding:18px 0}}@media(max-width:900px){{.metrics{{grid-template-columns:repeat(2,1fr)}}.grid{{grid-template-columns:1fr}}.shell{{padding:16px}}}}
</style></head><body><div class="shell"><header><div><h1>{title}</h1><div class="sub">Sistema automatizado | janela operacional: 01/04/2026 a {end:%d/%m/%Y} | exportação MT5: {rows[4][3]}</div></div><div><button onclick="document.documentElement.dataset.theme=document.documentElement.dataset.theme==='dark'?'light':'dark'">alternar tema</button><div class="sub" style="margin-top:7px">gerado em {now}</div></div></header><main><section class="metrics"><div class="metric"><label>P&L líquido</label><strong class="{'pos' if full['net']>=0 else 'neg'}">{money(full['net'])}</strong><span class="muted">{full['n']} trades</span></div><div class="metric"><label>Profit Factor</label><strong>{fbr(full['pf'],2)}</strong><span class="muted">acerto {pct(full['wr'])}</span></div><div class="metric"><label>Drawdown máximo</label><strong class="neg">{money(full['dd'])}</strong><span class="muted">{full['dur']} trades em DD</span></div><div class="metric"><label>Últimos 30 dias</label><strong class="{'pos' if recent30['net']>=0 else 'neg'}">{money(recent30['net'])}</strong><span class="muted">PF {fbr(recent30['pf'],2)}</span></div><div class="metric"><label>Última semana</label><strong class="{'pos' if recentwk['net']>=0 else 'neg'}">{money(recentwk['net'])}</strong><span class="muted">PF {fbr(recentwk['pf'],2)}</span></div></section><section class="grid"><div class="panel"><div class="panelhead">Curva de capital realizada</div><div class="body">{svg_line(eq,'P&L acumulado diário')}</div></div><div class="panel"><div class="panelhead">Leitura analítica</div><div class="body"><div class="callout {'good' if full['pf']>1 else 'bad'}"><h3>Carteira: edge frágil, mas positivo</h3><p>PF total de {fbr(full['pf'],2)} não sustenta aumento amplo de risco. A carteira depende de poucos motores positivos e carrega DD histórico de {money(full['dd'])}.</p></div><div class="callout bad"><h3>Risco prioritário</h3><p>RadarWIN e RadarWDO têm PF histórico abaixo de 1; o primeiro ainda tem associação alta com os demais robôs do WIN.</p></div><div class="callout"><h3>Concentração</h3><p>Maiores correlações diárias: {html.escape(strongtxt)}. O limite precisa ser conjunto, não definido por robô isolado.</p></div></div></div></section><section class="panel"><div class="panelhead">Desempenho por estratégia - desde o início operacional</div><div class="scroll"><table><thead>{tr(['Estratégia','Trades','P&L líquido','PF','Acerto','DD','Sharpe','Sortino','Decisão'],'th')}</thead><tbody>{perf}</tbody></table></div></section><section class="grid"><div class="panel"><div class="panelhead">Resultado mensal da carteira</div><div class="body">{svg_bars([x[0][5:] for x in monthly],[m(x[1])['net'] for x in monthly],'P&L mensal')}</div></div><div class="panel"><div class="panelhead">Drawdown realizado</div><div class="body">{svg_line(dd,'Drawdown acumulado','#e85d6c')}</div></div></section><section class="panel"><div class="panelhead">Subdivisões temporais</div><div class="scroll"><table><thead>{tr(['Estratégia','Desde 01/04','Últimos 30 dias','Semana corrente'],'th')}</thead><tbody>{temporal}</tbody></table></div></section><section class="panel"><div class="panelhead">Robustez e risco de cauda</div><div class="body"><p class="muted">Bootstrap IID: 2.000 reamostragens. Monte Carlo: 2.500 reordenações. Os intervalos não modelam mudança de regime, lote mínimo ou slippage.</p></div><div class="scroll"><table><thead>{tr(['Estratégia','IC 95% P&L','IC 95% expectativa','P(PF > 1)','DD p95 MC','P(DD > 1k)','P(DD > 2k)'],'th')}</thead><tbody>{robust}</tbody></table></div></section><section class="grid"><div class="panel"><div class="panelhead">Matriz de correlação diária de P&L</div><div class="scroll"><table><thead>{tr(['']+keys,'th')}</thead><tbody>{heat}</tbody></table></div></div><div class="panel"><div class="panelhead">Impacto marginal no drawdown</div><div class="body"><p class="muted">LOO: diferença entre o DD da carteira e o DD ao retirar a estratégia. Positivo significa redução de DD ao removê-la.</p></div><table><thead>{tr(['Estratégia','Redução de DD','DD sem estratégia'],'th')}</thead><tbody>{looh}</tbody></table></div></section><section class="panel"><div class="panelhead">Sensibilidade de sizing - modelo proporcional</div><div class="body"><p class="muted">RadarWIN, RadarWDO e Soberano são mantidos fixos. Os demais escalam proporcionalmente a partir de 1,6%. Não substitui reexecução com lote mínimo e stop individual.</p></div><table><thead>{tr(['Orçamento variável','P&L','PF','DD','Expectativa/trade'],'th')}</thead><tbody>{sizetable}</tbody></table></section><section class="panel"><div class="panelhead">Tabela mensal</div><div class="scroll"><table><thead>{tr(['Mês','Trades','P&L','PF','DD'],'th')}</thead><tbody>{monthlyh}</tbody></table></div></section><section class="panel"><div class="panelhead">Metodologia e limites</div><div class="body"><details open><summary>O que foi medido</summary><p>Posições realizadas foram associadas ao comentário da ordem de abertura. Símbolos foram normalizados para WIN, WDO e BIT. TurtleWIN começa em 26/06/2026 devido à configuração atual. As demais estratégias começam em 01/04/2026, início da janela operacional usada como referência.</p></details><details><summary>O que não pode ser concluído</summary><p>PF, bootstrap e Monte Carlo não validam uma estratégia isoladamente. A simulação de sizing não reproduz arredondamento por lote, margem, stops específicos, execução ou correlação intradiária.</p></details><details><summary>Recomendação operacional</summary><p>Não elevar o orçamento global enquanto a carteira estiver em semana negativa e os Radares permanecerem abaixo de PF 1. Manter WDO Turtle e WIN FVG como candidatos a continuidade; BIT Turtle requer confirmação; EngulfPattern permanece em sombra; Radares merecem redução ou pausa, salvo nova configuração comprovada e amostra adicional.</p></details></div></section></main><footer>Fonte: ReportHistory-2002705608.xlsx. Uso interno. Relatório quantitativo, não recomendação de investimento.</footer></div></body></html>'''
    curve=svg_multi({'Carteira':eq,**strategy_equity},'P&L acumulado diário por estratégia',days[0].strftime('%d/%m/%Y'),days[-1].strftime('%d/%m/%Y'))
    temporal_section=f'<section class="panel"><div class="panelhead">Subdivisões temporais</div><div class="body"><p class="muted">Métricas calculadas em cada janela. Sharpe e Sortino são por trade e não anualizados.</p>{temporal}</div></section>'
    analysis_html=f'''<section class="panel"><div class="panelhead">Análise dos resultados</div><div class="body"><div class="analysis"><h3>1. Qualidade do edge</h3><p>O portfólio encerra a janela com PF {fbr(full['pf'],2)} e P&L de {money(full['net'])}. Isso é positivo, mas estreito: o intervalo de robustez e o drawdown indicam que a margem de segurança estatística ainda é limitada. O resultado dos últimos 30 dias ({money(recent30['net'])}, PF {fbr(recent30['pf'],2)}) é favorável, porém a semana corrente ({money(recentwk['net'])}, PF {fbr(recentwk['pf'],2)}) mostra que a recuperação não deve ser extrapolada como novo regime.</p></div><div class="analysis"><h3>2. Motores e fragilidades</h3><p>As estratégias no topo da tabela por PF são as candidatas naturais a continuidade, desde que também sustentem a leitura rolling e a amostra. PF alto com poucos trades não é evidência suficiente: EngulfPattern deve permanecer em sombra. Em sentido oposto, RadarWIN e RadarWDO combinam PF histórico abaixo de 1 com risco de cauda material; não há justificativa estatística para aumentar o seu lote agora.</p></div><div class="analysis"><h3>3. Risco de carteira</h3><p>O DD histórico de {money(full['dd'])} ocorre numa carteira em que os robôs do WIN não são independentes. As correlações destacadas na matriz tornam Soberano, Turtle e RadarWIN um bloco de risco, não três apostas plenamente diversificadas. A retirada de cada estratégia no quadro LOO ajuda a distinguir contribuição de retorno de contribuição de drawdown, mas não substitui limites simultâneos por ativo.</p></div><div class="analysis"><h3>4. Decisão de sizing</h3><p>O teste proporcional mostra que elevar o orçamento amplia retorno e drawdown. Ele não deve ser usado como autorização automática porque não reproduz lotes inteiros, risco real por stop ou execução. A regra prudente é manter o orçamento atual até que a próxima revisão confirme PF de carteira acima de 1, expectativa não negativa e ausência de piora no DD; qualquer aumento deve respeitar um teto conjunto para os robôs WIN.</p></div></div></section>'''
    doc=doc.replace('</style>', '.dot{display:inline-block;width:8px;height:8px;margin-right:6px;vertical-align:middle}.legendbox{display:flex;flex-wrap:wrap;gap:8px 14px;margin:2px 0 8px}.legend{font:10px ui-monospace,monospace;color:var(--muted)}.legend i{display:inline-block;width:8px;height:8px;margin-right:5px}.period{margin-bottom:22px}.period h3{padding:5px 0;border-bottom:1px solid var(--border);color:var(--text)}.analysis{padding:12px 0;border-bottom:1px solid var(--border)}.analysis:last-child{border-bottom:0}</style>')
    doc=doc.replace('<thead><tr><th>Estratégia</th><th>Trades</th><th>P&L líquido</th><th>PF</th><th>Acerto</th><th>DD</th><th>Sharpe</th><th>Sortino</th><th>Decisão</th></tr></thead>', '<thead><tr><th>Estratégia</th><th>Início</th><th>Trades</th><th>P&L líquido</th><th>PF</th><th>Acerto</th><th>DD</th><th>Sharpe</th><th>Sortino</th><th>Decisão</th></tr></thead>')
    doc=re.sub(r'(<div class="panelhead">Curva de capital realizada</div><div class="body">).*?(</div></div><div class="panel"><div class="panelhead">Leitura analítica)', lambda x:x.group(1)+curve+x.group(2),doc,count=1,flags=re.DOTALL)
    doc=re.sub(r'<section class="panel"><div class="panelhead">Subdivisões temporais</div>.*?(?=<section class="panel"><div class="panelhead">Robustez e risco de cauda)', temporal_section, doc, count=1, flags=re.DOTALL)
    doc=doc.replace('Posições realizadas foram associadas ao comentário da ordem de abertura.', 'Posições não recebem estratégia sem um vínculo explícito no relatório; ordens e transações são classificadas apenas por seus próprios campos e vínculos explícitos.')
    doc = doc.replace(
        '</section></main><footer>',
        f'</section>{analysis_html}</main><footer>',
        1,
    )
    path = OUT / f"avaliacao_estrategias_{end:%Y-%m-%d}.html"
    path.write_text(doc, encoding="utf-8")
    print(f"Report: {title} -> {path}")


if __name__ == "__main__":
    main()
