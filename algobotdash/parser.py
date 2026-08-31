from __future__ import annotations

import math
from collections.abc import Iterable
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

from .config import ImportConfig

REPORT_TZ = ZoneInfo("America/Bahia")


@dataclass(frozen=True)
class PositionRecord:
    position_id: str
    entry_at: datetime
    exit_at: datetime | None
    symbol_raw: str
    symbol_family: str | None
    direction: str
    volume_requested: float | None
    volume_executed: float | None
    entry_price: float | None
    exit_price: float | None
    commission: float
    swap: float
    pnl: float
    comment: str
    strategy: str | None
    status: str


@dataclass(frozen=True)
class OrderRecord:
    order_id: str
    opened_at: datetime
    symbol_raw: str
    direction: str
    volume_requested: float | None
    volume_executed: float | None
    price: float | None
    stop_loss: float | None
    take_profit: float | None
    event_at: datetime | None
    status: str
    comment: str
    position_id: str | None
    strategy: str | None


@dataclass(frozen=True)
class TransactionRecord:
    transaction_id: str
    at: datetime
    symbol_raw: str
    direction: str
    volume: float | None
    price: float | None
    order_id: str | None
    commission: float
    tax: float
    swap: float
    pnl: float
    balance: float | None
    comment: str
    position_id: str | None
    strategy: str | None


@dataclass(frozen=True)
class RejectedRecord:
    row_number: int
    reason: str
    raw_position_id: str


def _cell(row: tuple[Any, ...], index: int) -> Any:
    return row[index] if index < len(row) else None


def _number(value: Any) -> float | None:
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    try:
        parsed = float(str(value).strip())
        return parsed if math.isfinite(parsed) else None
    except (TypeError, ValueError):
        return None


def _volume_pair(value: Any) -> tuple[float | None, float | None]:
    """Return requested and executed quantities without discarding either value."""
    if value is None or value == "":
        return None, None
    parts = [part.strip() for part in str(value).split("/")]
    if len(parts) == 1:
        amount = _number(parts[0])
        return amount, amount
    return _number(parts[0]), _number(parts[1])


def _datetime(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value if value.tzinfo else value.replace(tzinfo=REPORT_TZ)
    if not value:
        return None
    for fmt in ("%Y.%m.%d %H:%M:%S", "%Y.%m.%d %H:%M"):
        try:
            return datetime.strptime(str(value).strip(), fmt).replace(tzinfo=REPORT_TZ)
        except ValueError:
            continue
    return None


def _section_rows(rows: list[tuple[Any, ...]], section: str, next_section: str) -> Iterable[tuple[int, tuple[Any, ...]]]:
    start = next(i for i, row in enumerate(rows) if _cell(row, 0) == section)
    end = next((i for i in range(start + 1, len(rows)) if _cell(rows[i], 0) == next_section), len(rows))
    for index in range(start + 2, end):
        yield index + 1, rows[index]


def _last_section_rows(rows: list[tuple[Any, ...]], section: str, end_marker: str) -> Iterable[tuple[int, tuple[Any, ...]]]:
    start = next(i for i, row in enumerate(rows) if _cell(row, 0) == section)
    end = next((i for i in range(start + 1, len(rows)) if _cell(rows[i], 0) == end_marker), len(rows))
    for index in range(start + 2, end):
        yield index + 1, rows[index]


def _read_rows(source: str | Path) -> list[tuple[Any, ...]]:
    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        if sheet is None:
            raise ValueError(f"o workbook não possui planilha ativa: {source}")
        return [tuple(row) for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def _require_header(rows: list[tuple[Any, ...]], section: str, expected: tuple[tuple[int, Any], ...]) -> None:
    section_index = next((index for index, row in enumerate(rows) if _cell(row, 0) == section), None)
    if section_index is None or section_index + 1 >= len(rows):
        raise ValueError(f"a seção {section} não possui cabeçalho")
    header = rows[section_index + 1]
    if any(_cell(header, index) != value for index, value in expected):
        raise ValueError(f"cabeçalho inválido na seção {section}")


def _report_sections(rows: list[tuple[Any, ...]]) -> tuple[list[tuple[int, tuple[Any, ...]]], list[tuple[int, tuple[Any, ...]]], list[tuple[int, tuple[Any, ...]]]]:
    try:
        _require_header(rows, "Posições", ((0, "Horário"), (1, "Position"), (2, "Ativo"), (3, "Tipo"), (4, "Volume"), (5, "Preço"), (12, "Lucro")))
        _require_header(rows, "Ordens", ((0, "Horário da Abertura"), (1, "Ordem"), (2, "Ativo"), (3, "Tipo"), (4, "Volume"), (5, "Preço"), (8, "Horário"), (9, "Estado"), (11, "Comentário")))
        _require_header(rows, "Transações", ((0, "Horário"), (1, "Oferta"), (2, "Ativo"), (3, "Tipo"), (4, "Direção"), (5, "Volume"), (6, "Preço"), (7, "Ordem"), (11, "Lucro"), (12, "Saldo"), (13, "Comentário")))
        order_rows = list(_section_rows(rows, "Ordens", "Transações"))
        position_rows = list(_section_rows(rows, "Posições", "Ordens"))
        transaction_rows = list(_last_section_rows(rows, "Transações", "Posições Abertas"))
    except StopIteration as exc:
        raise ValueError("o workbook precisa conter as seções Posições, Ordens e Transações") from exc
    return position_rows, order_rows, transaction_rows


def _parse_positions(rows: Iterable[tuple[int, tuple[Any, ...]]], config: ImportConfig) -> tuple[list[PositionRecord], list[RejectedRecord]]:
    records: list[PositionRecord] = []
    rejected: list[RejectedRecord] = []
    for row_number, row in rows:
        position_id = str(_cell(row, 1) or "").strip()
        symbol = str(_cell(row, 2) or "").strip()
        if all(value in (None, "") for value in row):
            continue
        entry_at = _datetime(_cell(row, 0))
        pnl = _number(_cell(row, 12))
        if not position_id or not entry_at or not symbol or pnl is None:
            rejected.append(RejectedRecord(row_number, "campos obrigatórios inválidos", position_id))
            continue
        comment = ""
        strategy = None
        volume_requested, volume_executed = _volume_pair(_cell(row, 4))
        exit_at = _datetime(_cell(row, 8))
        records.append(PositionRecord(position_id, entry_at, exit_at, symbol, config.normalize_symbol(symbol), str(_cell(row, 3) or "").strip().lower(), volume_requested, volume_executed, _number(_cell(row, 5)), _number(_cell(row, 9)), _number(_cell(row, 10)) or 0.0, _number(_cell(row, 11)) or 0.0, pnl, comment, strategy, "closed" if exit_at else "open"))
    return records, rejected


def _parse_orders(rows: Iterable[tuple[int, tuple[Any, ...]]], config: ImportConfig) -> tuple[list[OrderRecord], dict[str, str | None], list[RejectedRecord]]:
    orders: list[OrderRecord] = []
    strategies: dict[str, str | None] = {}
    rejected: list[RejectedRecord] = []
    for row_number, row in rows:
        order_id = str(_cell(row, 1) or "").strip()
        opened_at = _datetime(_cell(row, 0))
        symbol = str(_cell(row, 2) or "").strip()
        if not order_id or not opened_at or not symbol:
            if any(value not in (None, "") for value in row):
                rejected.append(RejectedRecord(row_number, "campos obrigatórios inválidos em ordem", order_id))
            continue
        requested, executed = _volume_pair(_cell(row, 4))
        comment = str(_cell(row, 11) or "").strip()
        strategy = config.classify_strategy(comment)
        orders.append(OrderRecord(order_id, opened_at, symbol, str(_cell(row, 3) or "").strip().lower(), requested, executed, _number(_cell(row, 5)), _number(_cell(row, 6)), _number(_cell(row, 7)), _datetime(_cell(row, 8)), str(_cell(row, 9) or "").strip().lower(), comment, None, strategy))
        strategies[order_id] = strategy
    return orders, strategies, rejected


_TRANSACTION_NUMERIC_FIELDS = (5, 6, 8, 9, 10, 11, 12)


def _parse_transaction_numeric_values(row: tuple[Any, ...]) -> dict[int, float | None]:
    return {index: _number(_cell(row, index)) for index in _TRANSACTION_NUMERIC_FIELDS}


def _invalid_transaction_numeric_fields(
    row: tuple[Any, ...], numeric_values: dict[int, float | None]
) -> list[int]:
    return [
        index
        for index, parsed in numeric_values.items()
        if _cell(row, index) not in (None, "")
        and not (isinstance(_cell(row, index), str) and not _cell(row, index).strip())
        and parsed is None
    ]


def _parse_transactions(
    rows: Iterable[tuple[int, tuple[Any, ...]]], order_strategies: dict[str, str | None]
) -> tuple[list[TransactionRecord], list[RejectedRecord]]:
    transactions: list[TransactionRecord] = []
    rejected: list[RejectedRecord] = []
    for row_number, row in rows:
        transaction_id = str(_cell(row, 1) or "").strip()
        at = _datetime(_cell(row, 0))
        order_id = str(_cell(row, 7) or "").strip()
        if not transaction_id and not at:
            continue
        if not transaction_id or not at:
            if any(value not in (None, "") for value in row):
                rejected.append(RejectedRecord(row_number, "campos obrigatórios inválidos em transação", transaction_id))
            continue
        numeric_values = _parse_transaction_numeric_values(row)
        if _invalid_transaction_numeric_fields(row, numeric_values):
            rejected.append(RejectedRecord(row_number, "valores numéricos inválidos em transação", transaction_id))
            continue
        strategy = order_strategies.get(order_id)
        transactions.append(TransactionRecord(transaction_id, at, str(_cell(row, 2) or "").strip(), str(_cell(row, 4) or _cell(row, 3) or "").strip().lower(), numeric_values[5], numeric_values[6], order_id or None, numeric_values[8] if numeric_values[8] is not None else 0.0, numeric_values[9] if numeric_values[9] is not None else 0.0, numeric_values[10] if numeric_values[10] is not None else 0.0, numeric_values[11] if numeric_values[11] is not None else 0.0, numeric_values[12], str(_cell(row, 13) or "").strip(), None, strategy))
    return transactions, rejected


def read_report(source: str | Path, config: ImportConfig) -> tuple[list[PositionRecord], list[OrderRecord], list[TransactionRecord], list[RejectedRecord], int]:
    rows = _read_rows(source)
    position_rows, order_rows, transaction_rows = _report_sections(rows)
    positions, rejected_positions = _parse_positions(position_rows, config)
    orders, order_strategies, rejected_orders = _parse_orders(order_rows, config)
    transactions, rejected_transactions = _parse_transactions(transaction_rows, order_strategies)
    rejected = rejected_positions + rejected_orders + rejected_transactions
    return positions, orders, transactions, rejected, len(position_rows) + len(order_rows) + len(transaction_rows)


def read_positions(source: str | Path, config: ImportConfig) -> tuple[list[PositionRecord], list[RejectedRecord]]:
    """Compatibility helper for callers that only need the position summary."""
    records, _, _, rejected, _ = read_report(source, config)
    return records, rejected
